"""Flask server for the visual PDF field picker."""

import csv
import io
import json
import socket
import subprocess
import threading
import webbrowser
from pathlib import Path

import pypdfium2 as pdfium
from flask import Flask, jsonify, render_template, request, send_file
from openpyxl import load_workbook

from src.automations.fill_pdfs import (
    fill_certificate,
    font_mappings_keys,
    load_recipients,
)

# Render scale: how many pixels per PDF point
RENDER_SCALE = 2.0


def _numeric_font_weight(style):
    """Convert a font style name to a CSS-style numeric weight."""
    normalized = style.lower().replace(" ", "")
    mappings = (
        (("thin", "hairline"), 100),
        (("extralight", "ultralight"), 200),
        (("light",), 300),
        (("medium",), 500),
        (("semibold", "demibold"), 600),
        (("extrabold", "ultrabold"), 800),
        (("black", "heavy"), 900),
        (("bold",), 700),
    )
    for names, weight in mappings:
        if any(name in normalized for name in names):
            return weight
    return 400


def _local_font_catalog():
    """Return installed TrueType/OpenType families with numeric weights and paths."""
    try:
        result = subprocess.run(
            ["fc-list", "--format", "%{family[0]}\t%{style[0]}\t%{file}\n"],
            check=True,
            capture_output=True,
            text=True,
            timeout=10,
        )
    except (FileNotFoundError, subprocess.SubprocessError):
        return []

    fonts = {}
    for line in result.stdout.splitlines():
        parts = line.split("\t", 2)
        if len(parts) != 3:
            continue
        family, style, path = (part.strip() for part in parts)
        if not family or family.startswith(".") or Path(path).suffix.lower() not in {".ttf", ".otf"}:
            continue
        weight = _numeric_font_weight(style)
        fonts[(family, weight)] = {"family": family, "weight": weight, "style": style, "path": path}
    return sorted(fonts.values(), key=lambda item: (item["family"].lower(), item["weight"]))


def pixel_to_pdf_coords(click_x, click_y, render_w, render_h, pdf_w_pt, pdf_h_pt):
    """Convert pixel coordinates (top-left origin) to PDF coordinates (bottom-left origin)."""
    pdf_x = click_x * (pdf_w_pt / render_w)
    pdf_y = pdf_h_pt - (click_y * (pdf_h_pt / render_h))
    return round(pdf_x, 1), round(pdf_y, 1)


def render_pdf_page(pdf_path, scale=RENDER_SCALE):
    """Render the first page of a PDF to a PNG image.

    Returns:
        tuple: (png_bytes, render_width_px, render_height_px, pdf_width_pt, pdf_height_pt)
    """
    pdf = pdfium.PdfDocument(pdf_path)
    page = pdf[0]
    bitmap = page.render(scale=scale)
    pil_image = bitmap.to_pil()
    buf = io.BytesIO()
    pil_image.save(buf, format="PNG")
    png_bytes = buf.getvalue()
    pdf_w_pt = page.get_width()
    pdf_h_pt = page.get_height()
    return png_bytes, pil_image.width, pil_image.height, pdf_w_pt, pdf_h_pt


def _find_free_port():
    """Find an available port on localhost."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("127.0.0.1", 0))
        return s.getsockname()[1]


# Column names that should not appear as placeable certificate fields
_IGNORED_HEADERS = {"email", "e-mail", "email_address", "email address", "e mail", "mail"}


def _read_recipient_headers(recipients_path):
    """Read placeable field headers from CSV or XLSX recipient data."""
    if Path(recipients_path).suffix.lower() == ".xlsx":
        workbook = load_workbook(recipients_path, read_only=True, data_only=True)
        try:
            for worksheet in workbook.worksheets:
                for row in worksheet.iter_rows(values_only=True):
                    headers = [str(value).strip() if value is not None else "" for value in row]
                    normalized = {header.lower() for header in headers if header}
                    if normalized & {"name", "full name", "full_name", "recipient", "position", "role"}:
                        break
                else:
                    continue
                break
            else:
                headers = []
        finally:
            workbook.close()
    else:
        with open(recipients_path, encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            headers = next(reader, [])
    return [h.strip() for h in headers if h.strip() and h.strip().lower() not in _IGNORED_HEADERS]


def create_app(template_pdf, recipients_file=None, output_config=None, done_event=None):
    """Create and configure the Flask application."""
    template_folder = Path(__file__).resolve().parent / "templates"
    if not (template_folder / "picker.html").exists():
        # Standalone binaries may unpack Python modules without package data. Fall back to
        # the checked-out project so the visual picker still works from the repository.
        project_template_folder = Path.cwd() / "src" / "configure" / "templates"
        if (project_template_folder / "picker.html").exists():
            template_folder = project_template_folder
    app = Flask(__name__, template_folder=str(template_folder))

    # Pre-render template image and cache info
    png_bytes, render_w, render_h, pdf_w_pt, pdf_h_pt = render_pdf_page(template_pdf)

    initial_config = {}
    if output_config and Path(output_config).exists():
        try:
            initial_config = json.loads(Path(output_config).read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            initial_config = {}

    csv_headers = []
    sample_data = {}
    if recipients_file and Path(recipients_file).exists():
        csv_headers = _read_recipient_headers(recipients_file)
        try:
            recipients = load_recipients(recipients_file)
            for header in csv_headers:
                key = header.strip().lower()
                values = [str(row.get(key, "")).strip() for row in recipients]
                values = [value for value in values if value]
                if values:
                    sample_data[key] = max(values, key=len)
        except Exception:
            # Field placement remains usable when optional sample extraction fails.
            sample_data = {}

    @app.route("/")
    def index():
        return render_template("picker.html")

    @app.route("/api/template-image")
    def template_image():
        return send_file(io.BytesIO(png_bytes), mimetype="image/png")

    @app.route("/api/template-info")
    def template_info():
        return jsonify({
            "width_pt": pdf_w_pt,
            "height_pt": pdf_h_pt,
            "render_width_px": render_w,
            "render_height_px": render_h,
        })

    @app.route("/api/current-config")
    def get_current_config():
        return jsonify(initial_config)

    @app.route("/api/csv-headers")
    def get_csv_headers():
        return jsonify(csv_headers)

    @app.route("/api/sample-data")
    def get_sample_data():
        return jsonify(sample_data)

    @app.route("/api/font-families")
    def get_font_families():
        return jsonify(font_mappings_keys())

    @app.route("/api/local-fonts")
    def get_local_fonts():
        return jsonify(_local_font_catalog())

    @app.route("/api/preview", methods=["POST"])
    def preview():
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        config = {
            "font_family": data.get("font_family", "Helvetica"),
            "fields": data.get("fields", {}),
        }

        # Build sample recipient data from field names
        sample_data = {}
        for field_name in config["fields"]:
            sample_data[field_name] = data.get("sample_data", {}).get(field_name, field_name.title())

        import tempfile

        with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
            tmp_path = tmp.name

        try:
            fill_certificate(template_pdf, config, sample_data, tmp_path)
            preview_png, _, _, _, _ = render_pdf_page(tmp_path)
            return send_file(io.BytesIO(preview_png), mimetype="image/png")
        finally:
            Path(tmp_path).unlink(missing_ok=True)

    @app.route("/api/save", methods=["POST"])
    def save_config():
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data provided"}), 400

        config = {
            "template_pdf": str(template_pdf),
            "output_directory": data.get("output_directory", "local/outputs/fill-pdfs"),
            "font_family": data.get("font_family", "Helvetica"),
            "fields": data.get("fields", {}),
        }

        save_path = Path(output_config or "local/configs/fill-pdfs.json")
        save_path.parent.mkdir(parents=True, exist_ok=True)
        save_path.write_text(json.dumps(config, indent=2))

        if done_event:
            done_event.set()

        return jsonify({"saved": str(save_path)})

    return app


def launch_picker(template_pdf, recipients_file=None, output_config=None):
    """Launch the visual field picker in a browser.

    Args:
        template_pdf: Path to the PDF template file.
        recipients_file: Optional path to a CSV file (for column header extraction).
        output_config: Optional path for the output config JSON.

    Returns:
        str: The path to the saved configuration file.
    """
    template_pdf = str(Path(template_pdf).resolve())
    if not Path(template_pdf).exists():
        raise FileNotFoundError(f"Template PDF not found: {template_pdf}")

    done_event = threading.Event()
    port = _find_free_port()
    config_path = output_config or "local/configs/fill-pdfs.json"

    app = create_app(
        template_pdf=template_pdf,
        recipients_file=recipients_file,
        output_config=config_path,
        done_event=done_event,
    )

    def run_server():
        app.run(host="127.0.0.1", port=port, debug=False, use_reloader=False)

    server_thread = threading.Thread(target=run_server, daemon=True)
    server_thread.start()

    url = f"http://127.0.0.1:{port}"
    print(f"Opening visual picker at {url}")
    webbrowser.open(url)

    # Block until save is called
    done_event.wait()
    print(f"Configuration saved to {config_path}")
    return config_path
