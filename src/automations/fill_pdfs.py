import csv
import json
import os
import platform
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image, ImageDraw, ImageFont
from pypdf import PdfReader, PdfWriter
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas


def register_futura_font():
    """Register Futura font with ReportLab if available"""
    try:
        if platform.system() == "Darwin":  # macOS
            # Try to register Futura from system fonts
            futura_paths = [
                "/System/Library/Fonts/Supplemental/Futura.ttc",
                "/System/Library/Fonts/Futura.ttc",
                "/Library/Fonts/Futura.ttc",
            ]

            for font_path in futura_paths:
                if os.path.exists(font_path):
                    try:
                        # Register normal and bold variants with correct subfont indices
                        # Index 0: Futura-Medium (normal)
                        # Index 2: Futura-Bold (bold, non-italic)
                        pdfmetrics.registerFont(TTFont("Futura", font_path, subfontIndex=0))
                        pdfmetrics.registerFont(TTFont("Futura-Bold", font_path, subfontIndex=2))
                        print(f"Successfully registered Futura font from: {font_path}")
                        return True
                    except Exception as e:
                        print(f"Failed to register Futura from {font_path}: {e}")
                        continue

        # If we get here, Futura couldn't be registered
        print("Warning: Could not register Futura font. Using Helvetica as fallback.")
        return False

    except Exception as e:
        print(f"Error registering Futura font: {e}")
        return False


def _draw_rasterized_font_text(c, text, font_path, font_size, max_width, x, y, alignment, color):
    """Draw local CFF/OpenType text as a high-resolution transparent image."""
    scale = 4

    def load_font(size):
        return ImageFont.truetype(font_path, max(1, round(size * scale)))

    font = load_font(font_size)
    draw = ImageDraw.Draw(Image.new("RGBA", (1, 1)))
    measured_width = draw.textlength(text, font=font) / scale
    if max_width and measured_width > max_width:
        font_size = max(6, font_size * max_width / measured_width)
        font = load_font(font_size)

    bbox = draw.textbbox((0, 0), text, font=font, anchor="ls")
    left, top, right, bottom = bbox
    image = Image.new("RGBA", (max(1, right - left), max(1, bottom - top)), (0, 0, 0, 0))
    image_draw = ImageDraw.Draw(image)
    rgb = tuple(round(value * 255) for value in color) if all(0 <= value <= 1 for value in color) else tuple(color)
    image_draw.text((-left, -top), text, font=font, fill=(*rgb, 255), anchor="ls")

    width = image.width / scale
    height = image.height / scale
    draw_x = x - width / 2 if alignment == "center" else x - width if alignment == "right" else x
    draw_y = y - (bottom / scale)
    c.drawImage(ImageReader(image), draw_x, draw_y, width=width, height=height, mask="auto")
    return font_size


def load_recipients(recipients_file):
    """Load recipients from file (supports both single names and tab-separated format)"""
    recipients = []
    try:
        if recipients_file.lower().endswith(".xlsx"):
            workbook = load_workbook(recipients_file, read_only=True, data_only=True)
            try:
                for worksheet in workbook.worksheets:
                    rows = worksheet.iter_rows(values_only=True)
                    for row in rows:
                        headers = [str(value).strip() if value is not None else "" for value in row]
                        normalized_headers = [header.lower() for header in headers]
                        if any(
                            header in {"name", "full name", "full_name", "recipient", "position", "role"}
                            for header in normalized_headers
                        ):
                            break
                    else:
                        continue
                    break
                else:
                    raise ValueError("Could not find a header row in the XLSX file")

                for row in rows:
                    values = list(row)
                    if not any(value is not None and str(value).strip() for value in values):
                        continue
                    normalized = {
                        header: (str(values[index]).strip() if values[index] is not None else "")
                        for index, header in enumerate(normalized_headers)
                        if header and index < len(values)
                    }
                    raw_name = (
                        normalized.get("name")
                        or normalized.get("full name")
                        or normalized.get("full_name")
                        or normalized.get("recipient")
                    )
                    if not raw_name:
                        continue
                    normalized["name"] = raw_name
                    normalized["position"] = (
                        normalized.get("position")
                        or normalized.get("role")
                        or normalized.get("designation")
                        or ""
                    )
                    recipients.append(normalized)
            finally:
                workbook.close()
            return recipients

        # CSV support (header-based). Prefer by extension to avoid mis-detection.
        if recipients_file.lower().endswith(".csv"):
            with open(recipients_file, encoding="utf-8-sig", newline="") as f:
                reader = csv.DictReader(f)
                for row in reader:
                    # Normalize keys to lowercase and strip
                    normalized = {
                        str(k).strip().lower(): (v.strip() if isinstance(v, str) else v)
                        for k, v in row.items()
                        if k is not None
                    }

                    # Handle extra fields (DictReader puts them under key None)
                    extras = row.get(None) if isinstance(row, dict) else None
                    if isinstance(extras, list):
                        extras = [e.strip() for e in extras if isinstance(e, str)]

                    # Determine name with heuristics
                    raw_name = (
                        normalized.get("name")
                        or normalized.get("full_name")
                        or normalized.get("full name")  # Handle "Full Name" headers
                        or normalized.get("recipient")
                    )

                    # If the 'name' cell looks like an index (digits) and there are extras or a shifted layout,
                    # shift columns: treat current 'position' as name, current 'e-mail' as position (best-effort)
                    def is_index_like(x: str | None) -> bool:
                        return isinstance(x, str) and x.strip().isdigit()

                    if is_index_like(raw_name):
                        # If Position holds the real name, use it
                        possible_name = (
                            normalized.get("position")
                            or normalized.get("role")
                            or normalized.get("designation")
                        )
                        if possible_name:
                            raw_name = possible_name
                            # Shift position from email if present (common in misaligned CSV)
                            possible_pos = normalized.get("e-mail") or (
                                extras[0] if isinstance(extras, list) and extras else None
                            )
                            if possible_pos:
                                normalized["position"] = possible_pos

                    name = raw_name
                    if not name:
                        # Skip rows without a discernible name
                        continue
                    normalized["name"] = name
                    recipients.append(normalized)
            return recipients

        # TXT/TSV fallback (existing behavior)
        with open(recipients_file, encoding="utf-8") as f:
            for line_num, line in enumerate(f, 1):
                line = line.strip()
                # Skip empty lines and comments
                if not line or line.startswith("#"):
                    continue

                # Split by tab for multiple fields, else treat as single-name line
                parts = line.split("\t")
                if len(parts) >= 2:
                    recipients.append({"name": parts[0].strip(), "position": parts[1].strip()})
                elif len(parts) == 1:
                    recipients.append({"name": parts[0].strip(), "position": ""})
                else:
                    print(
                        f"Warning: Invalid format on line {line_num} in {recipients_file}. Expected format: Name or Name\tPosition"
                    )

    except FileNotFoundError:
        raise FileNotFoundError(f"Recipients file not found: {recipients_file}")
    except Exception as e:
        raise Exception(f"Error reading recipients file: {e!s}")

    return recipients


def load_config(config_file):
    """Load certificate configuration from JSON file"""
    try:
        with open(config_file, encoding="utf-8") as f:
            config = json.load(f)

        required_keys = ["template_pdf", "output_directory", "fields"]
        for key in required_keys:
            if key not in config:
                raise KeyError(f"Missing required configuration key: {key}")

        return config
    except FileNotFoundError:
        raise FileNotFoundError(f"Configuration file not found: {config_file}")
    except json.JSONDecodeError as e:
        raise Exception(f"Invalid JSON in configuration file: {e!s}")


def create_text_overlay(config, recipient_data, page_width, page_height):
    """Create a PDF overlay with text fields"""
    buffer = BytesIO()

    # Create canvas with the same page size as template
    c = canvas.Canvas(buffer, pagesize=(page_width, page_height))

    # Register Futura font if needed
    futura_available = register_futura_font()

    # Get font family from config
    font_family = config.get("font_family", "Helvetica")

    # Process each field
    for field_name, field_config in config["fields"].items():
        if recipient_data.get(field_name):
            text = recipient_data[field_name]
            x = field_config["x"]
            y = field_config["y"]
            font_size = field_config["font_size"]
            font_weight = field_config.get("font_weight", 400)
            numeric_weight = (
                700 if font_weight == "bold" else 400 if font_weight == "normal" else int(font_weight)
            )
            color = field_config.get("color", [0, 0, 0])
            alignment = field_config.get("alignment", "left")

            # Set font with improved fallback handling
            font_name = None

            # Define available standard fonts and their mappings
            font_mappings = {
                "Helvetica": {"normal": "Helvetica", "bold": "Helvetica-Bold"},
                "Times": {"normal": "Times-Roman", "bold": "Times-Bold"},
                "Times-Roman": {"normal": "Times-Roman", "bold": "Times-Bold"},
                "Courier": {"normal": "Courier", "bold": "Courier-Bold"},
                "Arial": {"normal": "Helvetica", "bold": "Helvetica-Bold"},
                "sans-serif": {"normal": "Helvetica", "bold": "Helvetica-Bold"},
                "serif": {"normal": "Times-Roman", "bold": "Times-Bold"},
                "monospace": {"normal": "Courier", "bold": "Courier-Bold"},
            }

            # Add Futura mapping if available
            if futura_available:
                font_mappings["Futura"] = {"normal": "Futura", "bold": "Futura-Bold"}
            else:
                font_mappings["Futura"] = {"normal": "Helvetica", "bold": "Helvetica-Bold"}

            # Get the appropriate font
            weight = "bold" if numeric_weight >= 600 else "normal"

            local_font_name = None
            raster_font_path = None
            font_path = field_config.get("font_path")
            if font_path and Path(font_path).is_file():
                try:
                    local_font_name = f"LocalFont-{abs(hash((font_path, numeric_weight)))}"
                    if local_font_name not in pdfmetrics.getRegisteredFontNames():
                        pdfmetrics.registerFont(TTFont(local_font_name, font_path))
                except Exception as error:
                    print(f"Note: Rasterizing local font '{font_path}' because PDF embedding failed: {error}")
                    local_font_name = None
                    raster_font_path = font_path

            if raster_font_path:
                fitted_size = _draw_rasterized_font_text(
                    c,
                    text,
                    raster_font_path,
                    font_size,
                    float(field_config["max_width"]) if field_config.get("max_width") else None,
                    x,
                    y,
                    alignment,
                    color,
                )
                if fitted_size != font_size:
                    print(
                        f"Note: Fitted '{field_name}' to max width {field_config['max_width']:g}pt "
                        f"at {fitted_size:g}pt: {text[:30]}{'...' if len(text) > 30 else ''}"
                    )
                continue

            if local_font_name:
                font_name = local_font_name
            elif font_family in font_mappings:
                font_name = font_mappings[font_family][weight]
            else:
                # Default fallback for unknown fonts
                font_name = "Helvetica-Bold" if numeric_weight >= 600 else "Helvetica"

            # Set the font with fallback
            try:
                c.setFont(font_name, font_size)
                if font_family == "Futura" and futura_available:
                    print("Note: Using registered Futura font")
                elif font_family == "Futura" and not futura_available:
                    print("Note: Using Helvetica as substitute for Futura font")
            except Exception:
                print(f"Warning: Font '{font_name}' not available, using Helvetica fallback")
                fallback_font = "Helvetica-Bold" if numeric_weight >= 600 else "Helvetica"
                font_name = fallback_font
                c.setFont(fallback_font, font_size)

            # Keep the historical length-based reduction when no explicit width is set.
            base_size = font_size
            max_width = field_config.get("max_width")
            if max_width:
                max_width = float(max_width)
                measured_width = c.stringWidth(text, font_name, base_size)
                if measured_width > max_width:
                    font_size = max(6, round(base_size * max_width / measured_width, 1))
                    while c.stringWidth(text, font_name, font_size) > max_width and font_size > 6:
                        font_size = round(font_size - 0.1, 1)
                    c.setFont(font_name, font_size)
                    print(
                        f"Note: Fitted '{field_name}' to max width {max_width:g}pt at {font_size:g}pt: "
                        f"{text[:30]}{'...' if len(text) > 30 else ''}"
                    )
            elif len(text) > 20:
                reduction = (len(text) - 20) * 0.4
                font_size = max(base_size * 0.5, base_size - reduction)
                font_size = round(font_size)
                if font_size != base_size:
                    c.setFont(font_name, font_size)
                    print(
                        f"Note: Reduced font size from {base_size} to {font_size}px for long text in '{field_name}': "
                        f"{text[:30]}{'...' if len(text) > 30 else ''}"
                    )

            # Set color (convert if needed)
            if all(isinstance(val, (int, float)) and 0 <= val <= 1 for val in color):
                # Already normalized (0-1)
                c.setFillColorRGB(color[0], color[1], color[2])
            else:
                # Convert from 0-255 to 0-1
                c.setFillColorRGB(color[0] / 255, color[1] / 255, color[2] / 255)

            # Handle text alignment
            if alignment == "center":
                text_width = c.stringWidth(text, font_name, font_size)
                x = x - (text_width / 2)
            elif alignment == "right":
                text_width = c.stringWidth(text, font_name, font_size)
                x = x - text_width

            # Draw text
            c.drawString(x, y, text)

    c.save()
    buffer.seek(0)
    return buffer


def fill_certificate(template_path, config, recipient_data, output_path):
    """Fill a certificate template with recipient data"""
    try:
        # Read template PDF
        with open(template_path, "rb") as template_file:
            template_reader = PdfReader(template_file)
            template_page = template_reader.pages[0]

            # Get page dimensions
            page_width = float(template_page.mediabox.width)
            page_height = float(template_page.mediabox.height)

            # Create text overlay
            overlay_buffer = create_text_overlay(config, recipient_data, page_width, page_height)
            overlay_reader = PdfReader(overlay_buffer)
            overlay_page = overlay_reader.pages[0]

            # Merge template and overlay
            template_page.merge_page(overlay_page)

            # Write output
            writer = PdfWriter()
            writer.add_page(template_page)

            with open(output_path, "wb") as output_file:
                writer.write(output_file)

        return True

    except Exception as e:
        raise Exception(f"Error filling certificate: {e!s}")


def generate_certificates(recipients_file, config_file, base_dir="data/certificates"):
    """
    Generate personalized certificates from template PDF

    Args:
        recipients_file: Path to CSV file with recipient data
        config_file: Path to configuration JSON file
        base_dir: Base directory for certificate files

    Returns:
        dict: Summary of generation results
    """

    # Helper to resolve file/dir paths sensibly across workspace vs base_dir
    def _resolve_path(p: str, base: str, is_dir: bool = False) -> str:
        """Resolve a path without incorrectly prefixing workspace-relative paths.

        Rules:
        - Absolute paths: return as-is
        - Paths starting with known project anchors (e.g., "workspace/", "templates/", "configs/")
          are treated as project-relative and returned as-is
        - Existing paths (relative to CWD): returned as-is
        - Otherwise: joined to base
        """
        if not p:
            return p
        anchors = ("workspace/", "templates/", "configs/", "src/", "docs/", "scripts/")
        if os.path.isabs(p) or p.startswith(anchors):
            return p
        # For directories that may not exist yet, rely on anchor check above and fall through to base join
        if not is_dir and os.path.exists(p):
            return p
        return os.path.join(base, p)

    # Resolve recipients and config paths
    recipients_file = _resolve_path(recipients_file, base_dir, is_dir=False)
    config_file = _resolve_path(config_file, base_dir, is_dir=False)

    # Load configuration and data
    config = load_config(config_file)
    recipients = load_recipients(recipients_file)

    if not recipients:
        print("No valid recipients found.")
        return {"total": 0, "generated": 0, "failed": 0, "errors": []}

    # Setup paths (respect project-root paths for template/output)
    template_path = _resolve_path(config["template_pdf"], base_dir, is_dir=False)
    output_dir = _resolve_path(config["output_directory"], base_dir, is_dir=True)

    # Validate template file
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template PDF not found: {template_path}")

    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)

    print(f"Loaded {len(recipients)} recipients")
    print(f"Template: {template_path}")
    print(f"Output directory: {output_dir}")

    # Initialize counters
    generated_count = 0
    failed_count = 0
    errors = []

    used_names = {}
    for i, recipient in enumerate(recipients, 1):
        try:
            print(f"\n[{i}/{len(recipients)}] Processing certificate for {recipient['name']}...")

            # Create safe base filename (optionally include position/role to reduce collisions)
            def _sanitize(s: str) -> str:
                s = "".join(c for c in s if c.isalnum() or c in (" ", "-", "_"))
                return s.rstrip().replace(" ", "_")

            safe_name = _sanitize(recipient["name"])
            position = recipient.get("position") or recipient.get("role") or ""
            base = f"{safe_name}" if not position else f"{safe_name}_{_sanitize(position)}"

            # Ensure unique filename (avoid overwrites)
            candidate = f"{base}_certificate.pdf"
            output_path = os.path.join(output_dir, candidate)
            idx = used_names.get(candidate, 0)
            while os.path.exists(output_path):
                idx += 1
                candidate = f"{base}_certificate_{idx}.pdf"
                output_path = os.path.join(output_dir, candidate)
            used_names[candidate] = idx

            # Fill certificate
            fill_certificate(template_path, config, recipient, output_path)

            generated_count += 1
            print(f"  ✓ Certificate generated: {os.path.basename(output_path)}")

        except Exception as e:
            error_msg = f"Failed to generate certificate for {recipient['name']}: {e!s}"
            print(f"  ✗ {error_msg}")
            errors.append(error_msg)
            failed_count += 1
            continue

    # Print summary
    print(f"\n{'=' * 50}")
    print("CERTIFICATE GENERATION SUMMARY")
    print(f"{'=' * 50}")
    print(f"Total recipients: {len(recipients)}")
    print(f"Successfully generated: {generated_count}")
    print(f"Failed: {failed_count}")
    print(f"Output directory: {output_dir}")

    if errors:
        print("\nErrors:")
        for error in errors:
            print(f"  - {error}")

    return {
        "total": len(recipients),
        "generated": generated_count,
        "failed": failed_count,
        "errors": errors,
        "output_directory": output_dir,
    }


def font_mappings_keys():
    """Return the list of available font family names for the picker UI."""
    return ["Helvetica", "Times", "Times-Roman", "Courier", "Arial", "Futura", "sans-serif", "serif", "monospace"]


def fill_certificates_from_file(
    recipients_file="recipients.txt", config_file="config.json", base_dir="data/certificates"
):
    """
    Convenience function to generate certificates using default file paths
    """
    return generate_certificates(recipients_file, config_file, base_dir)
